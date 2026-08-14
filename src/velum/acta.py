"""Acta de anonimización y tabla de equivalencias.

El acta es la prueba de diligencia: qué se hizo, sobre qué fichero, con qué
método y con qué resultado. No contiene ningún dato personal.

La tabla de equivalencias sí los contiene, y por eso va aparte y con el aviso
del artículo 4.5 del RGPD dentro: si se conserva, lo hecho es seudonimización
y el documento sigue siendo dato personal para quien tenga acceso a ella.
"""

from __future__ import annotations

import csv
import json
from datetime import datetime, timezone
from pathlib import Path

from .modelo import CATEGORIAS, TIPOS
from .motor import Resultado

AVISO_ARTICULO_4_5 = (
    "AVISO — Este fichero permite deshacer la sustitución. Conforme al artículo 4.5 "
    "del RGPD, mientras exista, el tratamiento es una SEUDONIMIZACIÓN y el documento "
    "sustituido sigue siendo dato personal para quien tenga acceso a esta tabla. "
    "Custódiese con las mismas medidas que el expediente original o destrúyase."
)

DESCARGO = (
    "Ninguna herramienta detecta el cien por cien de los datos personales. La revisión "
    "del documento antes de aportarlo, remitirlo o publicarlo corresponde al profesional, "
    "que conserva su deber de secreto y su responsabilidad bajo el RGPD y la LOPDGDD."
)


def _ahora() -> str:
    return datetime.now(timezone.utc).astimezone().isoformat(timespec="seconds")


def construir_acta(
    resultado: Resultado,
    *,
    fichero_origen: str,
    hash_origen: str,
    fichero_salida: str,
    modo: str,
    categorias: list[str],
    conservar_fechas: bool,
    version: str,
) -> dict:
    return {
        "herramienta": "VELUM by Abogado Virtual",
        "version": version,
        "metodo": "reglas con dígito de control y léxico jurídico en español; sin IA",
        "procesamiento": "íntegramente local — el documento no ha salido de este equipo",
        "fecha": _ahora(),
        "fichero_origen": fichero_origen,
        "hash_sha256_origen": hash_origen,
        "fichero_resultante": fichero_salida,
        "modo_de_sustitucion": modo,
        "categorias_tratadas": categorias,
        "fechas_conservadas": conservar_fechas,
        "datos_sustituidos": {
            "total": resultado.total,
            "por_tipo": resultado.recuento_por_tipo(),
            "por_categoria": resultado.recuento_por_categoria(),
            "entidades_distintas": len(resultado.registro.valores),
        },
        "articulo_9_rgpd_detectado": resultado.hay_articulo_9,
        "menciones_articulo_9_solo_avisadas": resultado.menciones_articulo_9,
        "datos_de_menores_detectados": resultado.hay_menores,
        "control_de_calidad": {
            "relectura_realizada": True,
            "posibles_residuos": len(resultado.residuos),
            "tipos_residuales": sorted({e.tipo for e in resultado.residuos}),
        },
        "advertencia": DESCARGO,
    }


def escribir_acta(acta: dict, destino: Path) -> Path:
    destino.parent.mkdir(parents=True, exist_ok=True)
    destino.write_text(json.dumps(acta, ensure_ascii=False, indent=2), encoding="utf-8")

    legible = destino.with_suffix(".md")
    legible.write_text(_acta_legible(acta), encoding="utf-8")
    return destino


def _acta_legible(acta: dict) -> str:
    datos = acta["datos_sustituidos"]
    lineas = [
        "# Acta de anonimización",
        "",
        f"**Herramienta.** {acta['herramienta']} v{acta['version']}",
        f"**Método.** {acta['metodo']}",
        f"**Procesamiento.** {acta['procesamiento']}",
        f"**Fecha.** {acta['fecha']}",
        "",
        "## Fichero",
        "",
        f"- Origen: `{acta['fichero_origen']}`",
        f"- Huella SHA-256 del original: `{acta['hash_sha256_origen']}`",
        f"- Resultado: `{acta['fichero_resultante']}`",
        f"- Modo de sustitución: {acta['modo_de_sustitucion']}",
        f"- Fechas conservadas: {'sí' if acta['fechas_conservadas'] else 'no'}",
        "",
        "## Datos sustituidos",
        "",
        f"- Total de apariciones: **{datos['total']}**",
        f"- Entidades distintas: **{datos['entidades_distintas']}**",
        "",
        "| Tipo | Apariciones |",
        "| --- | ---: |",
    ]
    for tipo, cuenta in datos["por_tipo"].items():
        descripcion = TIPOS[tipo].descripcion if tipo in TIPOS else tipo
        lineas.append(f"| {descripcion} ({tipo}) | {cuenta} |")

    lineas += [
        "",
        "## Categorías especiales",
        "",
        f"- Datos del artículo 9 del RGPD: "
        f"{'detectados' if acta['articulo_9_rgpd_detectado'] else 'no detectados'}",
        f"- Menciones a categoría especial señaladas sin sustituir: "
        f"{acta.get('menciones_articulo_9_solo_avisadas', 0)}",
        f"- Datos de personas menores de edad: "
        f"{'detectados' if acta['datos_de_menores_detectados'] else 'no detectados'}",
        "",
        "## Control de calidad",
        "",
        f"- Relectura del documento ya anonimizado: realizada",
        f"- Posibles residuos señalados: {acta['control_de_calidad']['posibles_residuos']}",
    ]
    if acta["control_de_calidad"]["tipos_residuales"]:
        lineas.append(
            f"- Tipos con residuo: {', '.join(acta['control_de_calidad']['tipos_residuales'])}"
        )

    lineas += ["", "---", "", acta["advertencia"], ""]
    return "\n".join(lineas)


def escribir_equivalencias(resultado: Resultado, destino: Path) -> Path:
    """Tabla reversible. XLSX si hay openpyxl; si no, CSV."""
    destino.parent.mkdir(parents=True, exist_ok=True)
    filas = resultado.equivalencias()

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
    except ImportError:
        return _escribir_equivalencias_csv(filas, destino.with_suffix(".csv"))

    destino = destino.with_suffix(".xlsx")
    libro = Workbook()
    hoja = libro.active
    hoja.title = "Equivalencias"

    hoja.append([AVISO_ARTICULO_4_5])
    hoja.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    celda_aviso = hoja.cell(row=1, column=1)
    celda_aviso.font = Font(bold=True, color="9C0006")
    celda_aviso.alignment = Alignment(wrap_text=True, vertical="top")
    hoja.row_dimensions[1].height = 60

    cabeceras = ["Etiqueta", "Tipo", "Categoría", "Valor(es) original(es)", "Apariciones"]
    hoja.append(cabeceras)
    for celda in hoja[2]:
        celda.font = Font(bold=True)

    for fila in filas:
        hoja.append(
            [
                fila["etiqueta"],
                fila["tipo"],
                CATEGORIAS.get(fila["categoria"], fila["categoria"]),
                fila["valores_originales"],
                int(fila["apariciones"] or 0),
            ]
        )

    for columna, ancho in zip("ABCDE", (22, 14, 42, 52, 12)):
        hoja.column_dimensions[columna].width = ancho

    libro.save(str(destino))
    return destino


def _escribir_equivalencias_csv(filas: list[dict[str, str]], destino: Path) -> Path:
    with open(destino, "w", encoding="utf-8-sig", newline="") as fichero:
        fichero.write(f"# {AVISO_ARTICULO_4_5}\n")
        escritor = csv.DictWriter(
            fichero,
            fieldnames=["etiqueta", "tipo", "categoria", "valores_originales", "apariciones"],
            delimiter=";",
        )
        escritor.writeheader()
        escritor.writerows(filas)
    return destino
